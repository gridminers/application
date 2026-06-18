"""
Management Command zum Import der Ist-Kosten aus einer Excel-Tabelle.

Aufruf:
    python manage.py import_real_costs pfad/zur/datei.xlsx
    python manage.py import_real_costs datei.xlsx --sheet "Ist-Kosten" --dry-run
"""
from __future__ import annotations

import re
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from core.models import Application

# Spaltenüberschrift (Excel)  ->  Modellfeld
COLUMN_MAP = {
    "PSP-Element": "psp_element",
    "Materialkosten": "real_material_costs",
    "Fremdleistungen": "real_external_services",
    "Eigenleistungen": "real_internal_services",
    "Ingenieurleistungen Dritte": "real_engineering_services",
    "Materialkostenzuschlag": "real_material_surcharge",
    "Investitionszuschläge": "real_investment_surcharge",
    "Gesamtkosten": "real_total_costs",
}

# Felder, die als Decimal eingelesen werden (alle außer PSP)
DECIMAL_FIELDS = {v for k, v in COLUMN_MAP.items() if k != "PSP-Element"}


def parse_euro(value) -> Decimal | None:
    """
    Wandelt einen Excel-Zellwert wie '2.800,00 €' oder 2800.0 in ein Decimal um.
    Gibt None zurück, wenn die Zelle leer ist.
    """
    if value is None:
        return None

    # Bereits numerisch (openpyxl liefert oft float/int)
    if isinstance(value, (int, float)):
        return Decimal(str(value))

    text = str(value).strip()
    if not text:
        return None

    # Währungssymbole und Leerzeichen (inkl. geschütztem Leerzeichen) entfernen
    text = text.replace("€", "").replace("\xa0", "").replace(" ", "")

    # Deutsches Zahlenformat: 28.362,57  ->  28362.57
    # Tausenderpunkte entfernen, Komma als Dezimaltrenner ersetzen
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    # sonst: englisches Format / reine Ganzzahl -> unverändert lassen

    # Negative Klammer-Notation o.ä. abfangen
    text = re.sub(r"[^\d.\-]", "", text)

    if not text or text in {"-", "."}:
        return None

    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def normalize_psp(value) -> str:
    """Normalisiert die PSP-Nummer (Leerzeichen weg, einheitlich vergleichbar)."""
    if value is None:
        return ""
    return str(value).strip()


class Command(BaseCommand):
    help = "Importiert Ist-Kosten aus einer Excel-Tabelle und ordnet sie per PSP-Element zu."

    def add_arguments(self, parser):
        parser.add_argument("excel_path", help="Pfad zur Excel-Datei (.xlsx)")
        parser.add_argument(
            "--sheet",
            default="Ist-Kosten",
            help='Name des Tabellenblatts (Standard: "Ist-Kosten")',
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Nur anzeigen, was passieren würde – keine Änderungen speichern.",
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            raise CommandError("openpyxl ist nicht installiert. Bitte: pip install openpyxl")

        path = options["excel_path"]
        sheet_name = options["sheet"]
        dry_run = options["dry_run"]

        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except FileNotFoundError:
            raise CommandError(f"Datei nicht gefunden: {path}")

        if sheet_name not in wb.sheetnames:
            raise CommandError(
                f'Tabellenblatt "{sheet_name}" nicht gefunden. '
                f"Verfügbar: {', '.join(wb.sheetnames)}"
            )

        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)

        # --- Kopfzeile lesen und Spaltenindizes bestimmen ---------------------
        try:
            header = next(rows)
        except StopIteration:
            raise CommandError("Das Tabellenblatt ist leer.")

        header_clean = [str(h).strip() if h is not None else "" for h in header]
        col_index = {}
        for excel_col, field in COLUMN_MAP.items():
            if excel_col in header_clean:
                col_index[field] = header_clean.index(excel_col)
            else:
                self.stdout.write(
                    self.style.WARNING(f'Spalte "{excel_col}" nicht gefunden – wird übersprungen.')
                )

        if "psp_element" not in col_index:
            raise CommandError('Pflichtspalte "PSP-Element" fehlt in der Kopfzeile.')

        # --- Datenzeilen verarbeiten -----------------------------------------
        updated, not_found, skipped, ambiguous = 0, 0, 0, 0

        with transaction.atomic():
            for row_nr, row in enumerate(rows, start=2):
                psp = normalize_psp(row[col_index["psp_element"]])
                if not psp:
                    skipped += 1
                    continue

                # Passenden Antrag suchen
                qs = Application.objects.filter(psp_element=psp)
                count = qs.count()

                if count == 0:
                    not_found += 1
                    self.stdout.write(
                        self.style.WARNING(f"Zeile {row_nr}: Kein Antrag mit PSP '{psp}' gefunden.")
                    )
                    continue
                if count > 1:
                    ambiguous += 1
                    self.stdout.write(
                        self.style.WARNING(
                            f"Zeile {row_nr}: PSP '{psp}' ist mehrdeutig "
                            f"({count} Treffer) – wird übersprungen."
                        )
                    )
                    continue

                application = qs.first()

                # Werte setzen
                for field, idx in col_index.items():
                    if field == "psp_element":
                        continue
                    raw = row[idx]
                    value = parse_euro(raw) if field in DECIMAL_FIELDS else raw
                    setattr(application, field, value)

                if dry_run:
                    self.stdout.write(
                        f"[DRY-RUN] Würde Antrag #{application.pk} "
                        f"(PSP {psp}) aktualisieren."
                    )
                else:
                    application.save(update_fields=list(DECIMAL_FIELDS) + ["updated_at"])
                    self.stdout.write(
                        self.style.SUCCESS(f"Zeile {row_nr}: Antrag #{application.pk} (PSP {psp}) aktualisiert.")
                    )
                updated += 1

            if dry_run:
                # Änderungen verwerfen
                transaction.set_rollback(True)

        # --- Zusammenfassung --------------------------------------------------
        self.stdout.write("")
        self.stdout.write(self.style.HTTP_INFO("===== Zusammenfassung ====="))
        self.stdout.write(f"  Aktualisiert : {updated}")
        self.stdout.write(f"  Nicht gefunden: {not_found}")
        self.stdout.write(f"  Mehrdeutig   : {ambiguous}")
        self.stdout.write(f"  Übersprungen : {skipped}")
        if dry_run:
            self.stdout.write(self.style.WARNING("DRY-RUN – es wurde nichts gespeichert."))
